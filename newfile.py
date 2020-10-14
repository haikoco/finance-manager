from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime

dt = datetime.now()
mmm = []

# Создать рабочую книгу в Excel:
#wb = Workbook()
wb = load_workbook(filename='maneger.xlsx')
sheet = wb.active
sheet.title = 'data'
q = len(sheet['C'])
print(q)
row = q + 1
f = 1
sheet['A'+str(1)] = 'Балланс: '
sheet['A'+str(row)] = '#'
sheet['B'+str(row)] = 'Дох/Рас'
sheet['C'+str(row)] = 'trans'
sheet['D'+str(row)] = '0'
sheet['E'+str(row)] = '0'
sheet['F'+str(row)] = 'date'
sheet['G'+str(row)] = 'time'
i=q
while True:
	i += 1
	qq = len(sheet['C'])
	qqq= qq+2
	#sheet['B'+str(1)] = str('=SUM('+'C'+str(3)+':'+'C'+str(qqq)+')')
	sheet['B'+str(1)] = str('=D'+str(qqq))
	ll = input('Дох/Рас: ')
	if ll == str('Д'):
		m = input('Доход: ')
		dox = input('От куда: ')
		z = str('+'+str(m))
		f +=1
		sum =str('='+'D'+str(f)+'+'+str(m))
		mmm.append([i, ll, z, sum, dox, dt.strftime("%d.%m.%Y"), dt.strftime("%I:%M:%S")])
		
	elif ll == str('Р'):
		m = input('Расход: ')
		dox = input('На что: ')
		z = str('-'+str(m))
		f +=1
		sum =str('='+'D'+str(f)+'-'+str(m))
		mmm.append([i, ll, z, sum, dox, dt.strftime("%d.%m.%Y"), dt.strftime("%I:%M:%S")])
	else:
		break
		
		
for item in mmm:
    row += 1
    sheet['A'+str(row)] = item[0]
    sheet['B'+str(row)] = item[1]
    sheet['C'+str(row)] = item[2]
    sheet['D'+str(row)] = item[3]
    sheet['E'+str(row)] = item[4]
    sheet['F'+str(row)] = item[5]
    sheet['G'+str(row)] = item[6]
    
filename = 'manager.xlsx'
wb.save(filename)